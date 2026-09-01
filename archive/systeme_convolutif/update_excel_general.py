# -*- coding: utf-8 -*-
"""
Met à jour le classeur Excel 'systeme_convolutif_spectral.xlsx' pour généraliser
le système convolutif à l'ensemble des rapports non typiques 1/k (k=3..12),
en utilisant les formules exactes validées dans spectral_engine.py :

  Somme A(k,n) = (alphaA(k)/2) * k^n - offsetA(k)
  Somme B(k,n) = (alphaB(k)/2) * k^n - offsetB(k)
  alphaA(k) = 2*(k^4-k^2+1)/(k-1) / k^3
  alphaB(k) = k*alphaA(k)
  offsetA(k) = k/(k-1)
  offsetB(k) = (k^7-k^6+k)/(k-1)

Actions effectuées :
  1. Complète tblConstantes (feuille Parametres) pour k=3..12.
  2. Complète tblRegles (feuille Parametres) avec la règle Digamma retenue
     par rapport (n=10), en cohérence avec CATALOGUE_DIGAMMA.
  3. Complète tblConvolution (feuille Convolution) avec les termes signés de
     la table de convolution formelle pour k=3..12, n=10, suites A et B.
  4. Ajoute une nouvelle feuille 'Systeme General' avec :
     - Le rappel des formules générales (fermées) valables pour tout n entier.
     - Un tableau récapitulatif k, n, Somme A, Somme B, Digamma retenu,
       Premier reconstruit, Rang, pour k=3..12.
"""

import sys
from fractions import Fraction as F

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

from spectral_engine import (
    coeff_A, coeff_B, offset_A, offset_B, sumA, sumB,
    termes_suite, candidats_digamma, CATALOGUE_DIGAMMA,
    reconstruire_premier,
)

SRC = "systeme_convolutif_spectral.xlsx"


def alpha_A(k):
    return 2 * coeff_A(k) / F(k) ** 3


def alpha_B(k):
    return 2 * coeff_B(k) / F(k) ** 3


K_RANGE = range(3, 13)  # k = 3..12
N_REF = 10


def find_header_row(ws, label, col=1, search_range=100):
    for r in range(1, search_range):
        if ws.cell(row=r, column=col).value == label:
            return r
    raise ValueError(f"Section '{label}' introuvable.")


def insert_rows_fix_merges(ws, idx, amount):
    """openpyxl.Worksheet.insert_rows ne décale pas les plages fusionnées
    (limitation connue). On le fait manuellement : toute plage fusionnée dont
    min_row >= idx est déplacée de +amount lignes vers le bas, faute de quoi
    l'ancienne fusion se retrouve figée au milieu des nouvelles données et
    openpyxl refuse ensuite d'écrire dans les cellules non-ancres qu'elle
    recouvre (silencieusement ignorées à l'écriture)."""
    old_ranges = list(ws.merged_cells.ranges)
    to_shift = [mr for mr in old_ranges if mr.min_row >= idx]
    for mr in to_shift:
        ws.unmerge_cells(str(mr))
    ws.insert_rows(idx, amount=amount)
    for mr in to_shift:
        new_min_row = mr.min_row + amount
        new_max_row = mr.max_row + amount
        ws.merge_cells(
            start_row=new_min_row, start_column=mr.min_col,
            end_row=new_max_row, end_column=mr.max_col,
        )


def update_constantes(ws):
    """Complète tblConstantes pour k=3..12. Localise dynamiquement l'en-tête
    (car update_regles peut avoir décalé les lignes vers le bas) et insère des
    lignes si nécessaire pour ne pas empiéter sur la section suivante."""
    section_row = find_header_row(ws, "CONSTANTES DES SOMMES FERMÉES")
    header_row = section_row + 1
    first_data_row = header_row + 1

    tbl = ws.tables["tblConstantes"]

    n_rows = len(list(K_RANGE))
    needed_last_row = first_data_row + n_rows - 1
    old_last_row = header_row + 4  # ancienne table : k=2,3,4 + ligne "Tout autre" (4 lignes)
    extra_rows = needed_last_row - old_last_row
    if extra_rows > 0:
        insert_rows_fix_merges(ws, old_last_row + 1, extra_rows)

    new_ref = f"A{header_row}:G{needed_last_row}"

    for i, k in enumerate(K_RANGE):
        r = first_data_row + i
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=float(alpha_A(k)))
        ws.cell(row=r, column=3, value=float(alpha_B(k)))
        ws.cell(row=r, column=4, value=float(offset_A(k)))
        ws.cell(row=r, column=5, value=float(offset_B(k)))
        ws.cell(row=r, column=6, value=f"1/{k}")
        ws.cell(row=r, column=7, value="Généralisé (formule fermée validée)")

    tbl.ref = new_ref


def update_regles(ws):
    """Complète tblRegles (A5:G...) avec une ligne par rapport 1/k, n=10.
    Insère des lignes si nécessaire pour ne pas empiéter sur la section
    'SORTIE DE LA RÈGLE SÉLECTIONNÉE' qui suit."""
    tbl = ws.tables["tblRegles"]
    header_row = 5  # row containing 'Rapport'/'k'/'n'/... (fixed, called before other inserts)
    first_data_row = header_row + 1  # 6

    n_rows = len(list(K_RANGE))
    needed_last_row = first_data_row + n_rows - 1
    old_last_row = 6  # dernière ligne actuelle de la table (une seule ligne 1/3)
    extra_rows = needed_last_row - old_last_row
    if extra_rows > 0:
        insert_rows_fix_merges(ws, old_last_row + 1, extra_rows)

    new_ref = f"A{header_row}:G{needed_last_row}"

    for i, k in enumerate(K_RANGE):
        r = first_data_row + i
        n = N_REF
        pos_rel, signe = CATALOGUE_DIGAMMA.get((k, n), (None, None))
        position = n + pos_rel if pos_rel is not None else None
        ws.cell(row=r, column=1, value=f"1/{k}")
        ws.cell(row=r, column=2, value=k)
        ws.cell(row=r, column=3, value=n)
        ws.cell(row=r, column=4, value=6)  # exposant de reconstruction : k^6 (constant)
        ws.cell(row=r, column=5, value=position)
        ws.cell(row=r, column=6, value=signe)
        statut = "Validé par formule fermée générale" if position is not None else "Essai-erreur requis"
        ws.cell(row=r, column=7, value=statut)

    tbl.ref = new_ref


def update_convolution(ws):
    """Complète tblConvolution (A5:I...) avec les termes signés pour k=3..12,
    n=10, suites A et B. Remplace intégralement les lignes de données."""
    tbl = ws.tables["tblConvolution"]
    header_row = 5
    first_data_row = header_row + 1  # 6

    rows = []
    for k in K_RANGE:
        for suite in ("A", "B"):
            for t in termes_suite(suite, N_REF):
                rows.append((f"1/{k}", suite, t.ordre, t.coeff1, t.exposant1,
                              t.coeff2, t.exposant2, t.valeur(k), t.expression()))

    last_row = first_data_row + len(rows) - 1
    new_ref = f"A{header_row}:I{last_row}"

    for i, row in enumerate(rows):
        r = first_data_row + i
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)

    tbl.ref = new_ref


def build_systeme_general_sheet(wb):
    if "Systeme General" in wb.sheetnames:
        del wb["Systeme General"]
    ws = wb.create_sheet("Systeme General")

    ws["A1"] = "Système convolutif généralisé — tous les rapports non typiques 1/k"
    ws["A2"] = (
        "Formules fermées exactes, valables pour tout k entier >= 3 et tout n "
        "entier (positif ou négatif). Validées contre les exemples numériques "
        "fournis pour k=3..7."
    )

    ws["A4"] = "FORMULES GÉNÉRALES (n entier, positif ou négatif)"
    ws["A5"] = "Somme suite A(k,n) = (alphaA(k)/2) * k^n − offsetA(k)"
    ws["A6"] = "Somme suite B(k,n) = (alphaB(k)/2) * k^n − offsetB(k)"
    ws["A7"] = "alphaA(k) = 2*(k^4 − k^2 + 1) / [(k−1) * k^3]"
    ws["A8"] = "alphaB(k) = k * alphaA(k)"
    ws["A9"] = "offsetA(k) = k / (k−1)"
    ws["A10"] = "offsetB(k) = (k^7 − k^6 + k) / (k−1)"
    ws["A11"] = (
        "Digamma(k,n,position,signe) = Somme A(k,n) + signe * k^position, "
        "position ∈ {n−3, n−2}, signe ∈ {+1,−1}"
    )
    ws["A12"] = "P_candidat(k,n) = (Somme B(k,n) − Digamma(k,n,position,signe)) / k^6"
    ws["A13"] = (
        "Le rapport 1/k est reconstruit en retenant le seul candidat pour lequel "
        "P_candidat est un entier premier (voir catalogue Digamma pour la levée "
        "d'ambiguïté quand plusieurs candidats premiers coexistent)."
    )

    header_row = 16
    headers = ["Rapport", "k", "n", "Somme A", "Somme B",
               "Position Digamma", "Signe Digamma", "Digamma calculé",
               "Premier reconstruit", "Rang du premier", "Statut"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)

    r = header_row + 1
    for k in K_RANGE:
        n = N_REF
        a = sumA(k, n)
        b = sumB(k, n)
        try:
            cand = reconstruire_premier(k, n)
            pos, signe, digamma, p_val, rang = (
                cand.position, cand.signe, cand.digamma_calcule,
                int(cand.p_candidat), cand.rang_premier,
            )
            statut = "Résolu (formule générale)"
        except ValueError as e:
            pos = signe = digamma = p_val = rang = None
            statut = f"Ambigu : {e}"

        ws.cell(row=r, column=1, value=f"1/{k}")
        ws.cell(row=r, column=2, value=k)
        ws.cell(row=r, column=3, value=n)
        ws.cell(row=r, column=4, value=int(a))
        ws.cell(row=r, column=5, value=int(b))
        ws.cell(row=r, column=6, value=pos)
        ws.cell(row=r, column=7, value=signe)
        ws.cell(row=r, column=8, value=int(digamma) if digamma is not None else None)
        ws.cell(row=r, column=9, value=p_val)
        ws.cell(row=r, column=10, value=rang)
        ws.cell(row=r, column=11, value=statut)
        r += 1

    last_row = r - 1
    tbl = Table(displayName="tblSystemeGeneral", ref=f"A{header_row}:K{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)

    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws["A" + str(last_row + 3)] = (
        "Note : la table de convolution formelle complète (termes signés "
        "c1*k^e1 + c2*k^e2) pour chaque rapport figure dans la feuille "
        "'Convolution' (table tblConvolution), désormais étendue à k=3..12."
    )


def main():
    wb = openpyxl.load_workbook(SRC)

    update_regles(wb["Parametres"])
    update_constantes(wb["Parametres"])
    update_convolution(wb["Convolution"])
    build_systeme_general_sheet(wb)

    out = SRC
    try:
        wb.save(out)
    except PermissionError:
        out = "systeme_convolutif_spectral_genere.xlsx"
        wb.save(out)
    print("Classeur mis à jour :", out)


if __name__ == "__main__":
    main()
