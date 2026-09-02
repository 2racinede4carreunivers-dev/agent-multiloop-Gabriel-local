"""Génère le classeur Excel universel des suites A/B non typiques."""
from __future__ import annotations

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).parent
ARCHIVE = ROOT / "archive" / "systeme_convolutif"
SOURCE = ARCHIVE / "systeme_convolutif_spectral_genere.xlsx"
FALLBACK_SOURCE = ARCHIVE / "systeme_convolutif_spectral.xlsx"
OUTPUT = ROOT / "systeme_convolutif_spectral_general.xlsx"
SHEET_NAME = "Reconstruction Universelle"

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")


def _style_title(cell) -> None:
    cell.fill = TITLE_FILL
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.alignment = Alignment(horizontal="center")


def _style_section(cell) -> None:
    cell.fill = SECTION_FILL
    cell.font = Font(bold=True)


def _set_formula(ws, row: int, label: str, formula: str, note: str = "") -> None:
    ws.cell(row, 1, label)
    ws.cell(row, 2, formula)
    ws.cell(row, 3, note)


def build_reconstruction_sheet(workbook) -> None:
    if SHEET_NAME in workbook.sheetnames:
        del workbook[SHEET_NAME]
    ws = workbook.create_sheet(SHEET_NAME)
    ws.merge_cells("A1:F1")
    ws["A1"] = "Reconstruction universelle des suites A et B - rapport non typique 1/k"
    _style_title(ws["A1"])

    ws.merge_cells("A2:F2")
    ws["A2"] = (
        "Saisir k >= 3 et des longueurs n >= 7. Toutes les formules sont algébriques "
        "et reconstruisent les coefficients depuis deux sommes, sans valeurs propres à un k."
    )
    ws["A2"].alignment = Alignment(wrap_text=True)

    ws["A4"] = "PARAMETRES A MODIFIER"
    _style_section(ws["A4"])
    inputs = [
        ("k (rapport 1/k)", 5),
        ("n1 : première somme", 10),
        ("n2 : deuxième somme distincte", 11),
        ("n cible (>= n1 pour convolution)", 14),
    ]
    for row, (label, value) in enumerate(inputs, start=5):
        ws.cell(row, 1, label)
        ws.cell(row, 2, value)
        ws.cell(row, 2).fill = INPUT_FILL
        ws.cell(row, 2).font = Font(bold=True)
    ws["C5"] = "Condition : k >= 3"
    ws["C6"] = "Conditions : n1, n2 >= 7 et n1 <> n2"

    ws["A11"] = "SOMMES DE DEPART (CONVOLUTION FORMELLE)"
    _style_section(ws["A11"])
    ws["A12"], ws["B12"], ws["C12"] = "Somme", "n1", "n2"
    _set_formula(
        ws, 13, "A",
        "=IF(OR($B$5<3,$B$6<7,$B$7<7),NA(),(1+1/$B$5+1/($B$5^3*($B$5-1)))*$B$5^$B$6-$B$5/($B$5-1))",
        "A(n) = (1 + 1/k + 1/(k^3*(k-1))) * k^n - k/(k-1)",
    )
    ws["C13"] = (
        "=IF(OR($B$5<3,$B$6<7,$B$7<7),NA(),"
        "(1+1/$B$5+1/($B$5^3*($B$5-1)))*$B$5^$B$7-$B$5/($B$5-1))"
    )
    _set_formula(
        ws, 14, "B",
        "=IF(OR($B$5<3,$B$6<7,$B$7<7),NA(),($B$5+1+1/($B$5^2*($B$5-1)))*$B$5^$B$6+($B$5^6-$B$5-$B$5^7)/($B$5-1))",
        "B(n) = (k + 1 + 1/(k^2*(k-1))) * k^n + (k^6-k-k^7)/(k-1)",
    )
    ws["C14"] = (
        "=IF(OR($B$5<3,$B$6<7,$B$7<7),NA(),"
        "($B$5+1+1/($B$5^2*($B$5-1)))*$B$5^$B$7+($B$5^6-$B$5-$B$5^7)/($B$5-1))"
    )

    ws["A17"] = "RECONSTRUCTION ALGEBRIQUE DES EQUATIONS A ET B"
    _style_section(ws["A17"])
    ws["A18"], ws["B18"], ws["C18"], ws["D18"] = "Suite", "Coefficient de k^n", "Constante", "Equation"
    _set_formula(ws, 19, "A", "=(B13-C13)/($B$5^$B$6-$B$5^$B$7)")
    ws["C19"] = "=B13-B19*$B$5^$B$6"
    ws["D19"] = '="A(n) = ("&TEXT(B19,"0.###############")&")*k^n + ("&TEXT(C19,"0.###############")&")"'
    _set_formula(ws, 20, "B", "=(B14-C14)/($B$5^$B$6-$B$5^$B$7)")
    ws["C20"] = "=B14-B20*$B$5^$B$6"
    ws["D20"] = '="B(n) = ("&TEXT(B20,"0.###############")&")*k^n + ("&TEXT(C20,"0.###############")&")"'
    ws["A22"] = "Les coefficients proviennent exclusivement des deux équations S(n1), S(n2)."

    ws["A24"] = "EVALUATION ET CONTROLE CONVOLUTIF"
    _style_section(ws["A24"])
    ws["A25"], ws["B25"], ws["C25"], ws["D25"] = (
        "Suite", "Somme reconstruite à n cible", "Somme par convolution", "Ecart",
    )
    _set_formula(ws, 26, "A", "=B19*$B$5^$B$8+C19")
    ws["C26"] = (
        "=IF($B$8<$B$6,NA(),$B$5^($B$8-$B$6)*B13+"
        "((1-$B$5)*C19)*(($B$5^($B$8-$B$6)-1)/($B$5-1)))"
    )
    ws["D26"] = "=B26-C26"
    _set_formula(ws, 27, "B", "=B20*$B$5^$B$8+C20")
    ws["C27"] = (
        "=IF($B$8<$B$6,NA(),$B$5^($B$8-$B$6)*B14+"
        "((1-$B$5)*C20)*(($B$5^($B$8-$B$6)-1)/($B$5-1)))"
    )
    ws["D27"] = "=B27-C27"
    ws["A29"] = "Convolution : S(n+1) = k*S(n) + (1-k)*constante."
    ws["A30"] = "Contrôle : les écarts doivent être nuls (à la précision Excel)."

    ws["A32"] = "RECONSTRUCTION DU PREMIER (DIGAMMA)"
    _style_section(ws["A32"])
    ws["A33"] = "Règle : Digamma = A(n) + signe*k^position ; P = (B(n)-Digamma)/k^6."
    ws["A34"], ws["B34"], ws["C34"], ws["D34"], ws["E34"] = (
        "Position", "Signe", "Digamma", "P candidat", "Entier ?",
    )
    candidate_row = 35
    for position_formula in ("$B$8-3", "$B$8-3", "$B$8-2", "$B$8-2"):
        sign = 1 if candidate_row in (35, 37) else -1
        ws.cell(candidate_row, 1, f"={position_formula}")
        ws.cell(candidate_row, 2, sign)
        ws.cell(candidate_row, 3, f"=B26+B{candidate_row}*$B$5^A{candidate_row}")
        ws.cell(candidate_row, 4, f"=(B27-C{candidate_row})/$B$5^6")
        ws.cell(candidate_row, 5, f'=IF(D{candidate_row}=INT(D{candidate_row}),"Oui","Non")')
        candidate_row += 1

    ws["A41"] = "REPLI REEL UNIVERSEL PAR PUISSANCES ET RACINES CARREES"
    _style_section(ws["A41"])
    ws["A42"] = (
        "Aucun tableau d'exemple n'est utilisé : Xi=RACINE((k^(i-1))^2+(k^i)^2). "
        "Cette voie est tentée après l'échec entier, pour tout k>=3 et n>=10."
    )
    ws["A43"], ws["B43"], ws["C43"], ws["D43"], ws["E43"], ws["F43"] = (
        "Coefficient A", "Coefficient B", "Radicande A", "Radicande B", "Somme A réelle", "Somme B réelle",
    )
    ws["A44"] = "=($B$5^($B$8-4)-1)/($B$5-1)+$B$5^($B$8-2)+$B$5^($B$8-1)"
    ws["B44"] = "=($B$5^5-1)/($B$5-1)+($B$5^($B$8-3)-$B$5^6)/($B$5-1)+$B$5^($B$8-1)+$B$5^$B$8"
    ws["C44"] = "=(1+$B$5^2)*A44^2"
    ws["D44"] = "=(1+$B$5^2)*B44^2"
    ws["E44"] = "=A44*SQRT(1+$B$5^2)"
    ws["F44"] = "=B44*SQRT(1+$B$5^2)"
    ws["A46"] = "HUIT BRANCHES REELLES UNIVERSELLES"
    _style_section(ws["A46"])
    headers = ("Pos. Digamma", "Signe", "Pos. Zêta", "Digamma réel", "Zêta réel", "P candidat", "Entier ?", "Primalité")
    for col, header in enumerate(headers, start=1):
        ws.cell(47, col, header)
    branches = ((-3, 1, 6), (-3, -1, 6), (-2, 1, 6), (-2, -1, 6),
                (-3, 1, 7), (-3, -1, 7), (-2, 1, 7), (-2, -1, 7))
    for row, (offset, sign, zeta_position) in enumerate(branches, start=48):
        ws.cell(row, 1, f"=$B$8{offset:+d}")
        ws.cell(row, 2, sign)
        ws.cell(row, 3, zeta_position)
        ws.cell(row, 4, f"=$E$44+B{row}*$B$5^(A{row}-1)*SQRT(1+$B$5^2)")
        ws.cell(row, 5, f"=$B$5^(C{row}-1)*SQRT(1+$B$5^2)")
        ws.cell(row, 6, f"=($F$44-D{row})/E{row}")
        ws.cell(row, 7, f'=IF(ABS(F{row}-ROUND(F{row},0))<=0.000000001,"Oui","Non")')
        ws.cell(row, 8, (
            f'=IF(G{row}<>"Oui","Non entier",IF(ROUND(F{row},0)=2,"Premier",'
            f'IF(ROUND(F{row},0)<2,"Composé",IF(ROUND(F{row},0)>1000000000000,'
            f'"À vérifier par Gabriel",IF(SUMPRODUCT(--(MOD(ROUND(F{row},0),'
            f'ROW(INDIRECT("2:"&INT(SQRT(ROUND(F{row},0))))))=0,"Premier","Composé")))))'
        ))
    ws["A57"] = (
        "Primalité Excel : test exact par diviseurs pour les candidats <= 10^12. "
        "Au-delà, Gabriel vérifie avant d'annoncer un premier."
    )

    for column, width in {"A": 34, "B": 25, "C": 28, "D": 50, "E": 15}.items():
        ws.column_dimensions[column].width = width
    for column in ("G", "H"):
        ws.column_dimensions[column].width = 18
    for row in (2, 22, 29, 30, 33, 42, 46, 57):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A5"


def main() -> None:
    source = SOURCE if SOURCE.exists() else FALLBACK_SOURCE
    if not source.exists():
        raise FileNotFoundError(f"Classeur source introuvable : {source}")
    output = OUTPUT
    try:
        copy2(source, output)
    except PermissionError:
        output = ROOT / "systeme_convolutif_spectral_general_methode_reelle.xlsx"
        copy2(source, output)
    workbook = load_workbook(output)
    build_reconstruction_sheet(workbook)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.save(output)
    print(f"Classeur créé : {output}")


if __name__ == "__main__":
    main()
